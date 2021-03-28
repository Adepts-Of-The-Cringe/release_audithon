import xlrd
from openpyxl import load_workbook
import MySQLdb
import sys
import re

def transliterate(name):
   slovar = {'а':'a','б':'b','в':'v','г':'g','д':'d','е':'e','ё':'e',
      'ж':'zh','з':'z','и':'i','й':'i','к':'k','л':'l','м':'m','н':'n',
      'о':'o','п':'p','р':'r','с':'s','т':'t','у':'u','ф':'f','х':'h',
      'ц':'c','ч':'cz','ш':'sh','щ':'scz','ъ':'','ы':'y','ь':'','э':'e',
      'ю':'u','я':'ja', 'А':'A','Б':'B','В':'V','Г':'G','Д':'D','Е':'E','Ё':'E',
      'Ж':'ZH','З':'Z','И':'I','Й':'I','К':'K','Л':'L','М':'M','Н':'N',
      'О':'O','П':'P','Р':'R','С':'S','Т':'T','У':'U','Ф':'F','Х':'H',
      'Ц':'C','Ч':'CZ','Ш':'SH','Щ':'SCH','Ъ':'','Ы':'y','Ь':'','Э':'E',
      'Ю':'U','Я':'YA',',':'','?':'',' ':'_','~':'','!':'','@':'','#':'',
      '$':'','%':'','^':'','&':'','*':'','(':'',')':'','-':'','=':'','+':'',
      ':':'',';':'','<':'','>':'','\'':'','"':'','\\':'','/':'','№':'',
      '[':'',']':'','{':'','}':'','ґ':'','ї':'', 'є':'','Ґ':'g','Ї':'i',
      'Є':'e', '—':'', ',':'.'}
   for key in slovar:
      name = name.replace(key, slovar[key])
   return name

def get_date(argument):
    return("" + str(argument.year) + "." + str(argument.month) + "." + str(argument.day))

def get_tipe(cell):
    print(cell)
    if ('ОФЗ-ПД' in cell):
        return ('OFZ-PD')
    if ('ОФЗ-ПК' in cell):
        return ('OFZ-PK')
    if ('ОФЗ-н' in cell):
        return ('OFZ-n')
    if ('БОФЗ' in cell):
        return ('BOFZ')
    if ('ГСО-ППС' in cell):
        return ('GSO-PPS')
    if ('ОФЗ-АД' in cell):
        return ('OFZ-AD')
    return ('UNKNOWN')

def xlsx_parse(database, argument):
   book = load_workbook(argument)
   sheet = book.active
   code = 'UNKNOWN'
   tipe = 'UNKNOWN'
   release_date = 'UNKNOWN'
   canceled_date = 'UNKNOWN'
   emission = 0.0
   canceled_in_fact = 0.0
   reserved = 0
   for r in range(5, 70):
       code = sheet.cell(r, 1).value
       tipe = get_tipe(sheet.cell(r, 2).value)
       release_date = get_date(sheet.cell(r, 3).value)
       canceled_date = get_date(sheet.cell(r, 4).value)
       emission = sheet.cell(r, 5).value
       canceled_in_fact = sheet.cell(r, 6).value
       reserved = 0
       values = (code, tipe, release_date, canceled_date, emission, canceled_in_fact)
       print(values)
       cursor.execute(query, values)
    return ('SUCCESS: File ' + argument + ' added to DB ')



# MAIN
database = MySQLdb.connect (host='localhost', user = 'tasian', passwd = '', db = 'amount_canceled_securities')
cursor = database.cursor()

try:
    query = """CREATE TABLE """ + sys.argv[1] + """ (code CHAR(9), type CHAR(15), release_date CHAR(10), canceled_date CHAR(10), emission FLOAT, canceled_in_fact FLOAT, reserved CHAR(25), id BIGINT NOT NULL AUTO_INCREMENT, PRIMARY KEY(id))"""
    cursor.execute(query) 
except MySQLdb._exceptions.OperationalError:
    print ("The table ", sys.argv[1], " already exists. File will be added to it")
query = """insert into """ + sys.argv[1] + """ (code, type, release_date,  canceled_date, emission, canceled_in_fact, reserved) VALUES (%s, %s, %s, %s, %s, %s,  0)"""
print(xlsx_parse(database, sys.argv[2]))
cursor.close()
database.commit()
database.close()

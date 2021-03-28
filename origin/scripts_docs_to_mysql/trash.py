import xlrd
from openpyxl import load_workbook
import MySQLdb
import sys
import re

def transliterate(name):
   slovar = {'а':'a','б':'b','в':'v','г':'g','д':'d','е':'e','ё':'e',
      'ж':'zh','з':'z','и':'i','й':'i','к':'k','л':'l','м':'m','н':'n',
      'о':'o','п':'p','р':'r','с':'s','т':'t','у':'u','ф':'f','х':'h',
      'ц':'c','ч':'cz','ш':'sh','щ':'scz','ъ':'','ы':'y','ь':'','э':'e',
      'ю':'u','я':'ja', 'А':'A','Б':'B','В':'V','Г':'G','Д':'D','Е':'E','Ё':'E',
      'Ж':'ZH','З':'Z','И':'I','Й':'I','К':'K','Л':'L','М':'M','Н':'N',
      'О':'O','П':'P','Р':'R','С':'S','Т':'T','У':'U','Ф':'F','Х':'H',
      'Ц':'C','Ч':'CZ','Ш':'SH','Щ':'SCH','Ъ':'','Ы':'y','Ь':'','Э':'E',
      'Ю':'U','Я':'YA',',':'','?':'',' ':'_','~':'','!':'','@':'','#':'',
      '$':'','%':'','^':'','&':'','*':'','(':'',')':'','-':'','=':'','+':'',
      ':':'',';':'','<':'','>':'','\'':'','"':'','\\':'','/':'','№':'',
      '[':'',']':'','{':'','}':'','ґ':'','ї':'', 'є':'','Ґ':'g','Ї':'i',
      'Є':'e', '—':'', ',':'.'}
   for key in slovar:
      name = name.replace(key, slovar[key])
   return name

def get_date(argument):
    return("" + str(argument.year) + "." + str(argument.month) + "." + str(argument.day))

def get_vipusk(r):
    if (r < 4):
        return ('UNKNOWN')
    if (r >= 4 and r <= 27):
        return ('36004RMFS')
    if (r >= 28 and r <= 49):
        return ('36005RMFS')
    if (r >= 50 and r <= 82):
        return ('36006RMFS')
    if (r >= 83 and r <= 110):
        return ('36007RMFS')
    if (r >= 111 and r <= 139):
        return ('36008RMFS')
    return ('UNKNOWN')

def get_data(argument):
    return("" + str(argument.year) + '.' + str(argument.month) + '.' + str(argument.day))

def xlsx_parse(database, argument):
    book = load_workbook(argument)
    sheet = book.active
    vipusk = 'UNKNOWN'
    number = 0
    date = 'UNKNOWN'
    period = 0
    percent_cupon = 0.00
    percent_amortization = 0.00
    nominal_roubles = 0.00
    razmer_cupona = 0.00
    pogashenie_nominala = 0.00
    vsego_viplat_po_abligacii = 0.00
    reserved = 0
    for r in range(1, sheet.max_row):
        if (type(sheet.cell(r, 2).value) == int):
            vipusk = get_vipusk(r)
            number = sheet.cell(r, 2).value
            date = get_date(sheet.cell(r, 3).value)
            period = sheet.cell(r, 4).value
            percent_cupon = 0 if (sheet.cell(r, 5).value == None) else float(sheet.cell(r, 5).value)
            percent_amortization = 0 if (sheet.cell(r, 6).value == None) else float(sheet.cell(r, 6).value)
            nominal_roubles = 0 if (sheet.cell(r, 7).value == None) else float(sheet.cell(r, 7).value)
            razmer_cupona = 0 if (sheet.cell(r, 8).value == None) else float(sheet.cell(r, 8).value)
            pogashenie_nominala = 0 if (sheet.cell(r, 9).value == None) else float(sheet.cell(r, 9).value)
            vsego_viplat_po_abligacii = 0 if (sheet.cell(r, 10).value == None) else float(sheet.cell(r, 10).value)
            reserved = 0
            values = (vipusk, number, date, period, percent_cupon, percent_amortization, nominal_roubles, razmer_cupona, pogashenie_nominala, vsego_viplat_po_abligacii)
            print(values)
            cursor.execute(query, values)
    return ('SUCCESS: File ' + argument + ' added to DB ')



# MAIN
database = MySQLdb.connect (host='localhost', user = 'tasian', passwd = '', db = 'amount_canceled_securities')
cursor = database.cursor()

try:
    query = """CREATE TABLE """ + sys.argv[1] + """ (vipusk CHAR(10), number INT, date CHAR(10), period INT, percent_cupon FLOAT, percent_amortization FLOAT, nominal_roubles FLOAT, razmer_cupona FLOAT, pogashenie_nominala FLOAT, vsego_viplat_po_abligacii FLOAT, reserved CHAR(25), id BIGINT NOT NULL AUTO_INCREMENT, PRIMARY KEY(id))"""
    cursor.execute(query) 
except MySQLdb._exceptions.OperationalError:
    print ("The table ", sys.argv[1], " already exists. File will be added to it")
query = """insert into """ + sys.argv[1] + """ (vipusk, number, date, period, percent_cupon, percent_amortization, nominal_roubles, razmer_cupona, pogashenie_nominala, vsego_viplat_po_abligacii, reserved) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s,  0)"""
print(xlsx_parse(database, sys.argv[2]))
cursor.close()
database.commit()
database.close()

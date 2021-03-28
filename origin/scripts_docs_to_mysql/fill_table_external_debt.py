import xlrd
from openpyxl import load_workbook
import MySQLdb
import sys
import re

def transliterate(name):
   slovar = {'а':'a','б':'b','в':'v','г':'g','д':'d','е':'e','ё':'e',
      'ж':'zh','з':'z','и':'i','й':'i','к':'k','л':'l','м':'m','н':'n',
      'о':'o','п':'p','р':'r','с':'s','т':'t','у':'u','ф':'f','х':'h',
      'ц':'c','ч':'cz','ш':'sh','щ':'scz','ъ':'','ы':'y','ь':'','э':'e',
      'ю':'u','я':'ja', 'А':'A','Б':'B','В':'V','Г':'G','Д':'D','Е':'E','Ё':'E',
      'Ж':'ZH','З':'Z','И':'I','Й':'I','К':'K','Л':'L','М':'M','Н':'N',
      'О':'O','П':'P','Р':'R','С':'S','Т':'T','У':'U','Ф':'F','Х':'H',
      'Ц':'C','Ч':'CZ','Ш':'SH','Щ':'SCH','Ъ':'','Ы':'y','Ь':'','Э':'E',
      'Ю':'U','Я':'YA',',':'','?':'',' ':'_','~':'','!':'','@':'','#':'',
      '$':'','%':'','^':'','&':'','*':'','(':'',')':'','-':'','=':'','+':'',
      ':':'',';':'','<':'','>':'','\'':'','"':'','\\':'','/':'','№':'',
      '[':'',']':'','{':'','}':'','ґ':'','ї':'', 'є':'','Ґ':'g','Ї':'i',
      'Є':'e', '—':'', ',':'.'}

   for key in slovar:
      name = name.replace(key, slovar[key])
   return name

def file_format(argument):
    x = 0
    for i in range(len(argument) - 1, 0, -1):
        if (argument[i] == '.'):
            x = i
            break
    if (x == 0):
       return ('ERROR')
    else:
       return argument[x:len(argument)]

def get_date(argument):
    pattern = r'(?:0?[1-9]|[12][0-9]|3[01])[.-](?:0?[1-9]|1[0-2])[.-](?:19[0-9][0-9]|20[012][0-9])'
    dates = re.findall(pattern, argument)

    print(dates)
    if(len(dates) == 1):
        for i in range(0, len(dates[0])):
            if (dates[0][i] == '.' or dates[0][i] == '-'):
                c = dates[0][i]
                x = (i + 1)
                if ((x + 7) > (len(dates[0]))):
                    return ('ERROR: Too short date')
                else:
                    month = dates[0][x:x + 2]
                    year = dates[0][x + 3:x + 7]
                    break
    else:
        return ('ERROR: More then 1 dates (wtf?)')

    return (year + c + month + c + '01')

def get_tipe(cell):
    if ('ОВГВЗ' in cell):
        return ('OVGVZ')
    if (('Париж' in cell) and ('реструктур' in cell)):
        return ('PARIS')
    if (('Париж' in cell) and ('не член' in cell)):
        return ('NPARIS')
    if (('облиг' in cell) and ('займам' in cell)):
        return ('OBZ00')
    if (('Государств' in cell) and ('внешний' in cell)):
        return ('ITOGO')
    if (('валют' in cell) and ('гарант' in cell)):
        return ('GOSGAR')
    if (('бывш' in cell) and ('СЭВ' in cell)):
        return ('OFSEV')
    if (('Коммерч' in cell) and ('СССР' in cell)):
        return ('USSR')
    if (('международн' in cell) and ('организац' in cell)):
        return ('INTNAT')
    if (('Прочая' in cell) and ('задолж' in cell)):
        return ('OTHER')
    return ('UNKNOWN')

def xls_parse(database, argument):
    book = xlrd.open_workbook(argument, encoding_override="-16BE")
    sheet = book.sheet_by_index(0)

    month = get_date(argument)
    if ('ERROR' in month):
        return ('ERROR: Wrong file name ' + argument)

    usd = '0'
    eur = '0'
    tipe = 'UNKNOWN'
    reserved = 0
    for r in range(0, sheet.nrows):
        for c in range(0, sheet.ncols):
            x = sheet.cell(r, c).value
            if (type(x) == float):
                y = sheet.cell(r, c + 1).value
                usd = x
                eur = y if (type(y) == float) else eur
                tipe = get_tipe(sheet.cell(r, 0).value)
                reserved = 0
                values = (month, usd, eur, tipe)
                print(values)
                cursor.execute(query, values)
                break
    return ('SUCCESS: File ' + argument + ' added to DB ')

def xlsx_parse(database, argument):
   book = load_workbook(argument)
   sheet = book.active

   month = get_date(argument)
   if ('ERROR' in month):
       return ('ERROR: Wrong file name ' + argument)

   usd = '0'
   eur = '0'
   tipe = 'UNKNOWN'
   reserved = 0
   for r in range(1, sheet.max_row):
       for c in range(1, sheet.max_column):
           x = sheet.cell(r, c).value
           if (type(x) == float):
               y = sheet.cell(r, c + 1).value
               usd = x
               eur = y if (type(y) == float) else eur
               tipe = get_tipe(sheet.cell(r, 1).value)
               reserved = 0
               values = (month, usd, eur, tipe)
               print(values)
               cursor.execute(query, values)
               break
   return ('SUCCESS: File ' + argument + ' added to DB ')



# MAIN
database = MySQLdb.connect (host='localhost', user = 'tasian', passwd = '', db = 'debt')
cursor = database.cursor()

try:
    query = """CREATE TABLE """ + sys.argv[1] + """ (month DATE, usd FLOAT, eur FLOAT, type CHAR(7), reserved CHAR(25), id BIGINT NOT NULL AUTO_INCREMENT, PRIMARY KEY(id))"""
    cursor.execute(query)
except MySQLdb._exceptions.OperationalError:
    print ("The table ", sys.argv[1], " already exists. File will be added to it")

query = """INSERT INTO """ + sys.argv[1] + """ (month, usd, eur, type, reserved) VALUES (%s, %s, %s, %s, 0)"""
for i in range(2, len(sys.argv)):
    ff = file_format(sys.argv[i])
    print (sys.argv[i])
    if (ff == '.xls'):
        print(xls_parse(database, sys.argv[i]))
    elif (ff == '.xlsx'):
        print(xlsx_parse(database, sys.argv[i]))
    else:
        print ('Error: Unsupported file format')


cursor.close()
database.commit()
database.close()

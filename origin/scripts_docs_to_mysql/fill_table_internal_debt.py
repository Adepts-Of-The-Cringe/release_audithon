import xlrd
from openpyxl import load_workbook
import MySQLdb
import sys
import re

def transliterate(name):
   slovar = {'а':'a','б':'b','в':'v','г':'g','д':'d','е':'e','ё':'e',
      'ж':'zh','з':'z','и':'i','й':'i','к':'k','л':'l','м':'m','н':'n',
      'о':'o','п':'p','р':'r','с':'s','т':'t','у':'u','ф':'f','х':'h',
      'ц':'c','ч':'cz','ш':'sh','щ':'scz','ъ':'','ы':'y','ь':'','э':'e',
      'ю':'u','я':'ja', 'А':'A','Б':'B','В':'V','Г':'G','Д':'D','Е':'E','Ё':'E',
      'Ж':'ZH','З':'Z','И':'I','Й':'I','К':'K','Л':'L','М':'M','Н':'N',
      'О':'O','П':'P','Р':'R','С':'S','Т':'T','У':'U','Ф':'F','Х':'H',
      'Ц':'C','Ч':'CZ','Ш':'SH','Щ':'SCH','Ъ':'','Ы':'y','Ь':'','Э':'E',
      'Ю':'U','Я':'YA',',':'','?':'',' ':'_','~':'','!':'','@':'','#':'',
      '$':'','%':'','^':'','&':'','*':'','(':'',')':'','-':'','=':'','+':'',
      ':':'',';':'','<':'','>':'','\'':'','"':'','\\':'','/':'','№':'',
      '[':'',']':'','{':'','}':'','ґ':'','ї':'', 'є':'','Ґ':'g','Ї':'i',
      'Є':'e', '—':'', ',':'.'}

   for key in slovar:
      name = name.replace(key, slovar[key])
   return name

def file_format(argument):
    x = 0
    for i in range(len(argument) - 1, 0, -1):
        if (argument[i] == '.'):
            x = i
            break
    if (x == 0):
       return ('ERROR')
    else:
       return argument[x:len(argument)]

def get_date(argument):
    pattern = r'(?:0?[1-9]|[12][0-9]|3[01])[.-](?:0?[1-9]|1[0-2])[.-](?:19[0-9][0-9]|20[012][0-9]|[0-9][0-9])'
    dates = re.findall(pattern, argument)

    if(len(dates) < 3):
        print(dates[0])
        for i in range(0, len(dates[0])):
            if (dates[0][i] == '.' or dates[0][i] == '-'):
                c = dates[0][i]
                x = (i + 1)
                if ((x + 5) > (len(dates[0]))):
                    return ('ERROR: Too short date')
                else:
                    month = dates[0][x:x + 2]
                    year = dates[0][x + 3:x + 5]
                break
    else:
        return ('ERROR: More then 2 dates (wtf?)')

    return (year + c + month + c + '01')

def get_tipe(cell):
    if ('ОФЗ-ПД' in cell):
        return ('OFZ-PD')
    if ('ОФЗ-АД' in cell):
        return ('OFZ-AD')
    if ('ОФЗ-ПК' in cell):
        return ('OFZ-PK')
    if ('ОФЗ-ИН' in cell):
        return ('OFZ-IN')
    if ('ОФЗ-н' in cell):
        return('OFZ-n')
    if ('ГСО-ППС' in cell):
        return ('GSO-PPS')
    if ('ГСО-ФПС' in cell):
        return ('GSO-FPS')
    if ('ОВОЗ' in cell):
        return ('OVOZ')
    if ('БОФЗ' in cell):
        return ('BOFZ')
    if ('Итого' in cell):
        return ('ITOGO')
    return ('UNKNOWN')

def xls_parse(database, argument):
    book = xlrd.open_workbook(argument, encoding_override="-16BE")
    sheet = book.sheet_by_index(0)

    month = get_date(argument)
    if ('ERROR' in month):
        return ('ERROR: Wrong date ' + argument)

    billion_roubles = '0'
    tipe = 'UNKNOWN'
    reserved = 0
    for r in range(0, sheet.nrows):
        x1 = sheet.cell(r, 1).value
        x2 = sheet.cell(r, 2).value
        if (type(x1) == float and type(x2) == str):
            billion_roubles = x1
            tipe = get_tipe(x2)
            reserved = 0
            values = (month, billion_roubles, tipe)
            print(values)
            cursor.execute(query, values)
            break
    return ('SUCCESS: File ' + argument + ' added to DB ')

def xlsx_parse(database, argument):
   book = load_workbook(argument)
   sheet = book.active

   month = get_date(argument)
   if ('ERROR' in month):
       print(month)
       return ('ERROR: Wrong date ' + argument)

   billion_roubles = '0'
   tipe = 'UNKNOWN'
   reserved = 0
   for r in range(1, sheet.max_row + 1):
       x2 = sheet.cell(r, 1).value
       x1 = sheet.cell(r, 2).value
       if (type(x1) == float and type(x2) == str):
           billion_roubles = x1
           tipe = get_tipe(x2)
           reserved = 0
           values = (month, billion_roubles, tipe)
           print(values)
           cursor.execute(query, values)
   return ('SUCCESS: File ' + argument + ' added to DB ')



# MAIN
database = MySQLdb.connect (host='localhost', user = 'tasian', passwd = '', db = 'debt')
cursor = database.cursor()

try:
    query = """CREATE TABLE """ + sys.argv[1] + """ (month DATE, billion_roubles FLOAT, type CHAR(7), reserved CHAR(25), id BIGINT NOT NULL AUTO_INCREMENT, PRIMARY KEY(id))"""
    cursor.execute(query)
except MySQLdb._exceptions.OperationalError:
    print ("The table ", sys.argv[1], " already exists. File will be added to it")

query = """INSERT INTO """ + sys.argv[1] + """ (month, billion_roubles, type, reserved) VALUES (%s, %s, %s, 0)"""
for i in range(2, len(sys.argv)):
    ff = file_format(sys.argv[i])
    print (sys.argv[i])
    if (ff == '.xls'):
        print(xls_parse(database, sys.argv[i]))
    elif (ff == '.xlsx'):
        print(xlsx_parse(database, sys.argv[i]))
    else:
        print ('Error: Unsupported file format')


cursor.close()
database.commit()
database.close()

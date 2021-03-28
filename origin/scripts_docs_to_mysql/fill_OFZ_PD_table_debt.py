import xlrd
from openpyxl import load_workbook
import MySQLdb
import sys
import re

def transliterate(name):
   slovar = {'а':'a','б':'b','в':'v','г':'g','д':'d','е':'e','ё':'e',
      'ж':'zh','з':'z','и':'i','й':'i','к':'k','л':'l','м':'m','н':'n',
      'о':'o','п':'p','р':'r','с':'s','т':'t','у':'u','ф':'f','х':'h',
      'ц':'c','ч':'cz','ш':'sh','щ':'scz','ъ':'','ы':'y','ь':'','э':'e',
      'ю':'u','я':'ja', 'А':'A','Б':'B','В':'V','Г':'G','Д':'D','Е':'E','Ё':'E',
      'Ж':'ZH','З':'Z','И':'I','Й':'I','К':'K','Л':'L','М':'M','Н':'N',
      'О':'O','П':'P','Р':'R','С':'S','Т':'T','У':'U','Ф':'F','Х':'H',
      'Ц':'C','Ч':'CZ','Ш':'SH','Щ':'SCH','Ъ':'','Ы':'y','Ь':'','Э':'E',
      'Ю':'U','Я':'YA',',':'','?':'',' ':'_','~':'','!':'','@':'','#':'',
      '$':'','%':'','^':'','&':'','*':'','(':'',')':'','-':'','=':'','+':'',
      ':':'',';':'','<':'','>':'','\'':'','"':'','\\':'','/':'','№':'',
      '[':'',']':'','{':'','}':'','ґ':'','ї':'', 'є':'','Ґ':'g','Ї':'i',
      'Є':'e', '—':'', ',':'.'}
   for key in slovar:
      name = name.replace(key, slovar[key])
   return name

def get_date(argument):
    return("" + str(argument.year) + "." + str(argument.month) + "." + str(argument.day))

def get_vipusk(r):
    if (r < 5):
        return ('UNKNOWN')
    if (r >= 5 and r <= 15):
        return ('25083RMFS')
    if (r >= 16 and r <= 25):
        return ('25084RMFS')
    if (r >= 26 and r <= 48):
        return ('26205RMFS')
    if (r >= 49 and r <= 81):
        return ('26207RMFS')
    if (r >= 82 and r <= 104):
        return ('26209RMFS')
    if (r >= 105 and r <= 121):
        return ('26210RMFS')
    if (r >= 122 and r <= 144):
        return ('26211RMFS')
    if (r >= 145 and r <= 177):
        return ('26211RMFS')
    if (r >= 178 and r <= 200):
        return ('26215RMFS')
    if (r >= 201 and r <= 215):
        return ('26217RMFS')
    if (r >= 216 and r <= 250):
        return ('26218RMFS')
    if (r >= 251 and r <= 274):
        return ('26219RMFS')
    if (r >= 275 and r <= 289):
        return ('26220RMFS')
    if (r >= 290 and r <= 324):
        return ('26221RMFS')
    if (r >= 325 and r <= 342):
        return ('26222RMFS')
    if (r >= 343 and r <= 357):
        return ('26223RMFS')
    if (r >= 358 and r <= 383):
        return ('26224RMFS')
    if (r >= 384 and r <= 419):
        return ('26225RMFS')
    if (r >= 420 and r <= 438):
        return ('26226RMFS')
    if (r >= 439 and r <= 452):
        return ('26227RMFS')
    if (r >= 453 and r <= 477):
        return ('26228RMFS')
    if (r >= 478 and r <= 493):
        return ('26229RMFS')
    if (r >= 494 and r <= 536):
        return ('26230RMFS')
    if (r >= 537 and r <= 589):
        return ('26231RMFS')
    if (r >= 590 and r <= 608):
        return ('26232RMFS')
    if (r >= 609 and r <= 642):
        return ('26233RMFS')
    if (r >= 643 and r <= 655):
        return ('26234RMFS')
    if (r >= 656 and r <= 679):
        return ('26235RMFS')
    if (r >= 680 and r <= 697):
        return ('26236RMFS')
    return ('UNKNOWN')

def get_data(argument):
    return("" + str(argument.year) + '.' + str(argument.month) + '.' + str(argument.day))

def xlsx_parse(database, argument):
    book = load_workbook(argument)
    sheet = book.active
    vipusk = 'UNKNOWN'
    number = 0
    date = 'UNKNOWN'
    period = 0
    percent_cupon = 0.00
    percent_amortization = 0.00
    nominal_roubles = 0.00
    razmer_cupona = 0.00
    pogashenie_nominala = 0.00
    vsego_viplat_po_abligacii = 0.00
    reserved = 0
    for r in range(1, sheet.max_row):
        if (type(sheet.cell(r, 2).value) == int):
            vipusk = get_vipusk(r)
            number = sheet.cell(r, 2).value
            date = get_date(sheet.cell(r, 3).value)
            period = sheet.cell(r, 4).value
            percent_cupon = sheet.cell(r, 5).value
            percent_amortization = 0 if (sheet.cell(r, 6).value == None) else float(sheet.cell(r, 6).value)
            nominal_roubles = sheet.cell(r, 7).value
            razmer_cupona = sheet.cell(r, 8).value
            pogashenie_nominala = sheet.cell(r, 9).value
            if (pogashenie_nominala == None):
                pogashenie_nominala = 0
            vsego_viplat_po_abligacii = sheet.cell(r, 10).value
            reserved = 0
            values = (vipusk, number, date, period, percent_cupon, percent_amortization, nominal_roubles, razmer_cupona, pogashenie_nominala, vsego_viplat_po_abligacii)
            print(values)
            cursor.execute(query, values)
    return ('SUCCESS: File ' + argument + ' added to DB ')



# MAIN
database = MySQLdb.connect (host='localhost', user = 'tasian', passwd = '', db = 'amount_canceled_securities')
cursor = database.cursor()

try:
    query = """CREATE TABLE """ + sys.argv[1] + """ (vipusk CHAR(10), number INT, date CHAR(10), period INT, percent_cupon FLOAT, percent_amortization FLOAT, nominal_roubles FLOAT, razmer_cupona FLOAT, pogashenie_nominala FLOAT, vsego_viplat_po_abligacii FLOAT, reserved CHAR(25), id BIGINT NOT NULL AUTO_INCREMENT, PRIMARY KEY(id))"""
    cursor.execute(query) 
except MySQLdb._exceptions.OperationalError:
    print ("The table ", sys.argv[1], " already exists. File will be added to it")
query = """insert into """ + sys.argv[1] + """ (vipusk, number, date, period, percent_cupon, percent_amortization, nominal_roubles, razmer_cupona, pogashenie_nominala, vsego_viplat_po_abligacii, reserved) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s,  0)"""
print(xlsx_parse(database, sys.argv[2]))
cursor.close()
database.commit()
database.close()
